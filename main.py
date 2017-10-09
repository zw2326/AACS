from openpyxl import load_workbook
from time import strftime
import code
import glob
import os
import re
import sys

# code.interact(local=locals())

env = {
    'workspacedir': 'workspace',
    'cachedir': os.path.join('workspace', 'cache'),
    'resultdir': os.path.join('workspace', 'result'),
    'mylog': os.path.join('workspace', 'mylog.log'),

    'sync': True,
    'filter': None,
    'debug': False,

    'symbols': []
}


class SymbolMeta(object):
    def __init__(self, symbol):
        self.symbol = symbol
        self.inputfiles = []
        self.i2s = {} # map inputfile to a list of 3 sheets
        self.i2d = {} # map inputfile to date

    def Insert(self, inputfile, date, sheets):
        self.inputfiles.append(inputfile)
        self.i2s[inputfile] = sheets
        self.i2d[inputfile] = date


class Sheet(object):
    def __init__(self, name, sheet):
        self.name = name
        self.sheet = sheet
        self.k2v = {} # map item name to value
        self.k2k = {} # map internal item name to original name in file


def PrepareDirs():
    if not os.path.exists(env['cachedir']):
        os.makedirs(env['cachedir'])
    if not os.path.exists(env['resultdir']):
        os.makedirs(env['resultdir'])

def ParseArgs():
    global env
    ptr = 1
    while ptr < len(sys.argv):
        if sys.argv[ptr] == '--debug':
            env['debug'] = True
        elif sys.argv[ptr] == '--no-sync':
            env['sync'] = False
        elif sys.argv[ptr] == '--filter':
            if ptr + 1 >= len(sys.argv):
                print '--filter must specify a filter.'
                exit(1)
            ptr += 1
            env['filter'] = sys.argv[ptr]
        elif sys.argv[ptr] == '--help':
            print '''
Usage: python main.py [OPTIONS] SYMBOL [SYMBOL2] ...
Script to scan and visualize financial statements.

  --debug                Print out debug messages.
  --filter FILTER        Specify a time filter. Syntax are:
                         ">YYYYQ[1-3]", ">YYYYK"
  --no-sync              Use cached statements only.
            '''
            exit(0)
        else:
            while ptr < len(sys.argv):
                env['symbols'].append(sys.argv[ptr].upper())
                ptr += 1
        ptr += 1

# Prepare statements for a symbol.
def PrepareInputfiles(symbol):
    if env['sync'] == True:
        Sync(symbol)
    inputfiles = Filter(symbol)
    return inputfiles

# Download all statements available for a symbol.
def Sync(symbol):
    pass

# Filter statements for a symbol by time.
def Filter(symbol):
    allfiles = [f for f in os.listdir(env['cachedir']) if re.match('^{0}-.*-10[QK].xlsx'.format(symbol), f)]
    selectedfiles = []
    rejectedfiles = []
    for file in sorted(allfiles):
        if True: ### TODO: do filtering
            selectedfiles.append(file)
        else:
            rejectedfiles.append(file)
    LogDebug('Selected: {0}'.format(', '.join(selectedfiles)))
    LogDebug('Rejected: {0}'.format(', '.join(rejectedfiles)))
    return map(lambda x: os.path.join(env['cachedir'], x), selectedfiles)

# Process all selected statements for a symbol.
def ProcessInputfiles(symbol, inputfiles):
    result = SymbolMeta(symbol)

    for inputfile in inputfiles:
        LogDebug('  Processing file {0}'.format(inputfile))
        wb = load_workbook(filename = inputfile)

        ### TODO: Get sheet "Document and Entity Information" and extract
        ### Document period end date as SymbolMeta's i2d value.
        date = '2017-03-04'
        LogDebug('    Date set to: {0}'.format(date))

        # Sheet name is truncated to 31 characters. Use
        # the A1 element of each sheet as its full name.
        fullname2sheet = {}
        for name in wb.get_sheet_names():
            sheet = wb[name]
            fullname = sheet['A1'].value.lower()
            fullname2sheet[fullname] = sheet

        cbssheet = LocateCBS(wb, fullname2sheet)
        cbssheet = ProcessCBS(cbssheet)

        # Add (inputfile, date, [3 sheets]) to the meta data object for this symbol.
        result.Insert(inputfile, date, [cbssheet])
    return result

# Locate the consolidated balance sheet.
def LocateCBS(wb, fullname2sheet):
    candidates = []
    for fullname in fullname2sheet.keys():
        if re.match('^.*consolidated balance sheet.*$', fullname) and not re.match('^.*parenthetical.*$', fullname):
            candidates.append(fullname)
    if len(candidates) == 0:
        Error('We cannot find the consolidated balance sheet.')
    elif len(candidates) > 1:
        Error('We find more than one consolidated balance sheets: {0}'.format(', '.join(candidates)))
    LogDebug('    Found CBS: {0}'.format(candidates[0]))
    return Sheet(candidates[0], fullname2sheet[candidates[0]])

# Process consolidated balance sheet.
def ProcessCBS(cbssheet):
    for row in cbssheet.sheet.iter_rows(row_offset=1):
        if row[0].value == None:
            continue
        rowkey = row[0].value.lower()

        ### TODO: remove $, ' '; convert e.g. '9,821' to float
        ### TODO: save original row key
        # Assets.
        if re.match('^.*total current assets.*$', rowkey):
            cbssheet.k2v['total current assets'] = row[1].value
            cbssheet.k2k['total current assets'] = rowkey
        elif re.match('^.*total non-current assets.*$', rowkey):
            cbssheet.k2v['total non-current assets'] = row[1].value
            cbssheet.k2k['total non-current assets'] = rowkey
        elif re.match('^.*total assets.*$', rowkey):
            cbssheet.k2v['total assets'] = row[1].value
            cbssheet.k2k['total assets'] = rowkey
        # Liabilities.
        elif re.match('^.*total current liabilities.*$', rowkey):
            cbssheet.k2v['total current liabilities'] = row[1].value
            cbssheet.k2k['total current liabilities'] = rowkey
        elif re.match('^.*total non-current liabilities.*$', rowkey):
            cbssheet.k2v['total non-current liabilities'] = row[1].value
            cbssheet.k2k['total non-current liabilities'] = rowkey
        elif re.match('^.*total liabilities.*$', rowkey) and not re.match('^.*equity.*$', rowkey):
            cbssheet.k2v['total liabilities'] = row[1].value
            cbssheet.k2k['total liabilities'] = rowkey
        # Equity.
        elif re.match('^.*total.*equity.*$', rowkey) and not re.match('^.*liabilities.*$', rowkey):
            cbssheet.k2v['total equity'] = row[1].value
            cbssheet.k2k['total equity'] = rowkey
        # Liabilities and equity.
        elif re.match('^.*total liabilities and.*equity.*$', rowkey):
            cbssheet.k2v['total liabilities and equity'] = row[1].value
            cbssheet.k2k['total liabilities and equity'] = rowkey

    LogDebug('    Extracted following items from CBS:\n' + '\n'.join(['{0}({1}): {2}'.format(x, cbssheet.k2k[x].encode('utf-8'), cbssheet.k2v[x]) for x in cbssheet.k2v.keys()]))
    return cbssheet

# Generate HTML.
def Render(symbolmetas):
    outputfile = os.path.join(env['resultdir'], 'index.html')
    fid = open(outputfile, 'w')
    fid.write('''
<!DOCTYPE html>
<html>
<head>
<meta http-equiv="X-UA-Compatible" content="IE=edge">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/0.2.0/Chart.min.js" type="text/javascript"></script>
</head>
<body>
<canvas id="myChart" width="1600" height="800"></canvas>
<script>
    ''')

    for symbolmeta in symbolmetas:
        cbsdate = [symbolmeta.i2d[i] for i in symbolmeta.inputfiles]
        cbsdata = [symbolmeta.i2s[i][0] for i in symbolmeta.inputfiles]
        fid.write('''
var data = {
  labels: [
''' + ', '.join(map(lambda x: '"{0}"'.format(x), cbsdate)) + '''
],
  datasets: [
      {
          label: "Sugar intake",
          fillColor: "rgba(151,187,205,0.2)",
          strokeColor: "rgba(151,187,205,1)",
          pointColor: "rgba(151,187,205,1)",
          pointStrokeColor: "#fff",
          pointHighlightFill: "#fff",
          pointHighlightStroke: "rgba(151,187,205,1)",
          data: [
''' + ', '.join([str(x.k2v['total liabilities and equity']) for x in cbsdata]) + '''
]
      }
  ]
};

new Chart(document.getElementById("myChart").getContext("2d")).Line(data);
</script>
</body>
</html>
    ''')
    LogDebug('HTML {0} generated'.format(outputfile))
    fid.close()

def Log(msg):
    fid = open(env['mylog'], 'a')
    fid.write('[{0}] {1}\n'.format(strftime('%Y-%m-%d %H:%M:%S'), msg))
    fid.close()
    print msg

def LogDebug(msg):
    if env['debug'] == True:
        Log('DEBUG: ' + msg)

def Error(msg):
    Log('ERROR: ' + msg)
    exit(1)


if __name__ == '__main__':
    PrepareDirs()
    ParseArgs()

    symbolmetas = [] # [SymbolMeta1, ...]
    for symbol in env['symbols']:
        LogDebug('Processing symbol {0}'.format(symbol))
        inputfiles = PrepareInputfiles(symbol)
        symbolmeta = ProcessInputfiles(symbol, inputfiles)
        symbolmetas.append(symbolmeta)

    Render(symbolmetas)
